from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook


DEFAULT_RULES_PATH = Path("selection_rules.xlsx")


@dataclass
class SelectionRules:
    min_price: float
    preferred_price_min: float
    preferred_price_max: float
    max_review_count_for_recommend: int
    max_review_count_for_watch: int
    max_sponsored_in_top_n: int
    top_n_window: int
    require_bought_info_for_recommend: bool
    blocked_brands: list[str]
    risky_title_keywords: list[str]


def ensure_rules_workbook(path: str | Path) -> Path:
    workbook_path = Path(path)
    if workbook_path.exists():
        return workbook_path.resolve()

    workbook = Workbook()
    rules_sheet = workbook.active
    rules_sheet.title = "Rules"
    rules_sheet.append(["rule_key", "value", "description"])
    rules_sheet.append(["min_price", 10, "Low-price items below this are usually filtered out"])
    rules_sheet.append(["preferred_price_min", 15, "Prices at or above this level are considered healthier"])
    rules_sheet.append(["preferred_price_max", 40, "Prices within preferred min/max get extra credit"])
    rules_sheet.append(["max_review_count_for_recommend", 1000, "Above this is too competitive for recommend"])
    rules_sheet.append(["max_review_count_for_watch", 3000, "Above this is too competitive even for watch"])
    rules_sheet.append(["max_sponsored_in_top_n", 5, "If top-N results contain more than this many ads, competition is heavy"])
    rules_sheet.append(["top_n_window", 10, "Use the first N search results for ad-density checks"])
    rules_sheet.append(["require_bought_info_for_recommend", "TRUE", "Recommend only if bought-info exists"])

    brand_sheet = workbook.create_sheet("BlockedBrands")
    brand_sheet.append(["brand", "note"])
    brand_sheet.append(["Nike", "Strong brand / likely not suitable"])
    brand_sheet.append(["Apple", "Strong brand / likely not suitable"])
    brand_sheet.append(["Disney", "Trademark risk"])
    brand_sheet.append(["Pokemon", "Trademark risk"])

    keyword_sheet = workbook.create_sheet("RiskKeywords")
    keyword_sheet.append(["keyword", "note"])
    keyword_sheet.append(["medical", "Regulated or sensitive"])
    keyword_sheet.append(["fda", "Regulated or sensitive"])
    keyword_sheet.append(["baby", "Sensitive category"])
    keyword_sheet.append(["kids", "Sensitive category"])

    instruction_sheet = workbook.create_sheet("Instructions")
    instruction_sheet.append(["使用说明"])
    instruction_sheet.append(["1. 只改黄色高亮工作表中的 value 或列表内容，不要改表头。"])
    instruction_sheet.append(["2. Rules 工作表控制阈值。"])
    instruction_sheet.append(["3. BlockedBrands 写你不想碰的品牌。"])
    instruction_sheet.append(["4. RiskKeywords 写你想避开的标题关键词。"])
    instruction_sheet.append(["5. 程序会输出 decision、decision_reason、rule_* 等列。"])

    for sheet_name in ["Rules", "BlockedBrands", "RiskKeywords"]:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = _yellow_fill()

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    return workbook_path.resolve()


def _yellow_fill():
    from openpyxl.styles import PatternFill

    return PatternFill(fill_type="solid", fgColor="FFF2CC")


def load_rules(path: str | Path) -> SelectionRules:
    workbook_path = ensure_rules_workbook(path)
    workbook = load_workbook(workbook_path, data_only=True)

    rules_map: dict[str, object] = {}
    rules_sheet = workbook["Rules"]
    for row in rules_sheet.iter_rows(min_row=2, values_only=True):
        key, value, _description = row
        if key:
            rules_map[str(key).strip()] = value

    blocked_brands = _load_single_column_values(workbook["BlockedBrands"])
    risky_title_keywords = _load_single_column_values(workbook["RiskKeywords"])

    return SelectionRules(
        min_price=float(rules_map.get("min_price", 10)),
        preferred_price_min=float(rules_map.get("preferred_price_min", 15)),
        preferred_price_max=float(rules_map.get("preferred_price_max", 40)),
        max_review_count_for_recommend=int(rules_map.get("max_review_count_for_recommend", 1000)),
        max_review_count_for_watch=int(rules_map.get("max_review_count_for_watch", 3000)),
        max_sponsored_in_top_n=int(rules_map.get("max_sponsored_in_top_n", 5)),
        top_n_window=int(rules_map.get("top_n_window", 10)),
        require_bought_info_for_recommend=_to_bool(rules_map.get("require_bought_info_for_recommend", True)),
        blocked_brands=blocked_brands,
        risky_title_keywords=risky_title_keywords,
    )


def _load_single_column_values(sheet) -> list[str]:
    values: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[0]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            values.append(text)
    return values


def _to_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def annotate_products(rows: Iterable[dict], rules: SelectionRules) -> list[dict]:
    row_list = [dict(row) for row in rows]
    top_window = row_list[: max(rules.top_n_window, 1)]
    sponsored_count = sum(1 for row in top_window if _is_true(row.get("is_sponsored")))

    annotated: list[dict] = []
    for row in row_list:
        decision, reason, fields = evaluate_product(row, rules, sponsored_count)
        annotated_row = dict(row)
        annotated_row.update(fields)
        annotated_row["decision"] = decision
        annotated_row["decision_reason"] = reason
        annotated.append(annotated_row)
    return annotated


def evaluate_product(row: dict, rules: SelectionRules, sponsored_count: int) -> tuple[str, str, dict]:
    price = _to_float(row.get("price"))
    review_count = _to_int(row.get("review_count"))
    brand = str(row.get("brand") or "").strip()
    title = str(row.get("title") or "").strip()
    bought_info = str(row.get("bought_info") or "").strip()

    rule_price_ok = price is not None and price >= rules.min_price
    rule_preferred_price = (
        price is not None and rules.preferred_price_min <= price <= rules.preferred_price_max
    )
    rule_has_demand_signal = bool(bought_info)
    rule_review_ok_for_recommend = review_count is not None and review_count <= rules.max_review_count_for_recommend
    rule_review_ok_for_watch = review_count is not None and review_count <= rules.max_review_count_for_watch
    rule_brand_risk = _matches_any(brand, rules.blocked_brands)
    rule_title_risk = _matches_any(title, rules.risky_title_keywords)
    rule_competition_heavy = sponsored_count > rules.max_sponsored_in_top_n

    reasons: list[str] = []
    if not rule_price_ok:
        reasons.append("price below minimum")
    if rule_brand_risk:
        reasons.append("blocked brand")
    if rule_title_risk:
        reasons.append("risky title keyword")
    if rule_competition_heavy:
        reasons.append("heavy sponsored competition")

    if (
        rule_price_ok
        and rule_preferred_price
        and rule_review_ok_for_recommend
        and not rule_brand_risk
        and not rule_title_risk
        and not rule_competition_heavy
        and (rule_has_demand_signal or not rules.require_bought_info_for_recommend)
    ):
        decision = "recommend"
        reasons.insert(0, "healthy price and manageable competition")
        if rule_has_demand_signal:
            reasons.insert(1, "has bought signal")
    elif (
        rule_price_ok
        and rule_review_ok_for_watch
        and not rule_brand_risk
        and not rule_title_risk
    ):
        decision = "watch"
        if rule_has_demand_signal:
            reasons.insert(0, "has demand signal but needs manual review")
        else:
            reasons.insert(0, "passes baseline but demand signal is weak")
    else:
        decision = "skip"
        if not reasons:
            reasons.append("does not meet baseline rules")

    fields = {
        "rule_price_ok": rule_price_ok,
        "rule_preferred_price": rule_preferred_price,
        "rule_has_demand_signal": rule_has_demand_signal,
        "rule_review_ok_for_recommend": rule_review_ok_for_recommend,
        "rule_review_ok_for_watch": rule_review_ok_for_watch,
        "rule_brand_risk": rule_brand_risk,
        "rule_title_risk": rule_title_risk,
        "rule_competition_heavy": rule_competition_heavy,
        "top_window_sponsored_count": sponsored_count,
    }
    return decision, "; ".join(reasons), fields


def _matches_any(text: str, keywords: Iterable[str]) -> bool:
    lower_text = text.lower()
    return any(keyword.lower() in lower_text for keyword in keywords if keyword)


def _is_true(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() == "true"


def _to_float(value: object) -> Optional[float]:
    if value in (None, ""):
        return None
    return float(value)


def _to_int(value: object) -> Optional[int]:
    if value in (None, ""):
        return None
    return int(float(value))
